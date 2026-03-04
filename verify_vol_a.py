"""
verify_vol_a.py  –  Fix unmatched Vol A questions only.

For each DB question that has NO current match:
  1. Look for a sheet whose name == question number
     -> if found: check text matches DB text -> use it
  2. If no QN sheet or text mismatch: scan all sheets for matching text
     -> if found: add override
  3. If nothing found: leave as-is (tool will show TOC / available questions)

Writes new overrides to data/vol_a_overrides.json  (merges with existing).

Run:  python verify_vol_a.py [wave_filter]
  e.g.  python verify_vol_a.py 103
        python verify_vol_a.py          (all waves)
"""

import os, re, sys, json, unicodedata
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

DATA_DIR       = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
VOL_A_DIR      = os.path.join(DATA_DIR, "vol_a")
DB_PATH        = os.path.join(DATA_DIR, "enes.duckdb")
OVERRIDES_FILE = os.path.join(DATA_DIR, "vol_a_overrides.json")
QN_INDEX_FILE  = os.path.join(DATA_DIR, "vol_a_question_index.json")

WAVE_FILTER = sys.argv[1].strip().lower() if len(sys.argv) > 1 else None

import duckdb
try:
    import openpyxl; _XLSX = True
except ImportError:
    _XLSX = False; print("WARNING: openpyxl not installed")
try:
    import xlrd; _XLS = True
except ImportError:
    _XLS = False; print("WARNING: xlrd not installed")

_FRENCH_RE = re.compile(r'[àâçéèêëîïôùûüæœÀÂÇÉÈÊËÎÏÔÙÛÜÆŒ]')

# ── normalisation ──────────────────────────────────────────────────────────────
def norm_wave(w):
    return re.sub(r'^[Ee][Bb]\s*', '', w.strip()).strip()

def norm_qcode(s):
    return re.sub(r'[\s._-]', '', s).upper()

def norm_exact(text):
    if not text: return ''
    # Strip question codes like 'QA1.', 'QB7.1.', 'QC6_1.', 'QB7_2.'
    t = re.sub(r'^\s*[A-Z]{1,3}\d+[a-z]?([._]\d+[a-z]?)?\s*[\.\-]?\s*', '', text, count=1)
    # Strip survey piping/programming markers like *?PipeinUnderStart, *?Pipin_ST0595
    t = re.sub(r'\*\?[A-Za-z0-9_()]+', '', t)
    # Strip inline instructions (SHOW SCREEN, READ OUT, etc.)
    t = re.sub(r'\n.*', '', t)  # keep only first line
    return re.sub(r'\s+', ' ', t).strip().lower()

def texts_match(db, xl):
    a, b = norm_exact(db), norm_exact(xl)
    if not a or not b: return False
    if a == b: return True
    if len(a) >= 30 and b.startswith(a): return True
    if len(b) >= 30 and a.startswith(b): return True
    # Prefix match: texts differ only at suffix (e.g. '(M)' vs '(multiple answers possible)')
    min_len = min(len(a), len(b))
    if min_len >= 70 and a[:70] == b[:70]: return True
    return False

def wave_from_file(fname):
    m = re.search(r'[Ee][Bb](\d+\.\d+)', fname)
    if m: return m.group(1)
    m = re.match(r'eb(\d{3,4})_', fname, re.IGNORECASE)
    if m:
        d = m.group(1); return f"{d[:-1]}.{d[-1]}"
    m = re.search(r'\((\d+\.\d+)\)', fname)
    if m: return m.group(1)
    return None

# ── Excel helpers ──────────────────────────────────────────────────────────────
def get_sheets(fpath):
    try:
        if fpath.lower().endswith('.xlsx'):
            if not _XLSX: return []
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            n = list(wb.sheetnames); wb.close(); return n
        else:
            if not _XLS: return []
            wb = xlrd.open_workbook(fpath, on_demand=True)
            n = wb.sheet_names(); wb.release_resources(); return n
    except Exception as e:
        print(f"  [WARN] cannot open {os.path.basename(fpath)}: {e}"); return []

def get_sheet_text(fpath, sname):
    """Extract English question text from a sheet.

    Priority: cell whose first word matches the sheet name (e.g. 'QA1. text...')
    Fallback:  first long English cell that is not a survey title/header line.
    """
    sn_norm = norm_qcode(sname)
    _HEADER_SKIP = re.compile(
        r'^(eurobarometer|vol\s*a|base:|terrain|fieldwork|<<|ue\d|eu\d)',
        re.IGNORECASE
    )
    try:
        if fpath.lower().endswith('.xlsx'):
            if not _XLSX: return ''
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            if sname not in wb.sheetnames: wb.close(); return ''
            ws = wb[sname]
            primary = ''
            fallback = ''
            for row in ws.iter_rows(values_only=True, max_row=8):
                for cell in row:
                    if cell and isinstance(cell, str):
                        s = cell.strip()
                        if not s or len(s) <= 20 or _FRENCH_RE.search(s):
                            continue
                        fw = s.split()[0].rstrip('.,;:!?')
                        if norm_qcode(fw) == sn_norm:
                            primary = s; break
                        if not fallback and not _HEADER_SKIP.match(s):
                            fallback = s
                if primary: break
            wb.close()
            return primary or fallback
        else:
            if not _XLS: return ''
            wb = xlrd.open_workbook(fpath, on_demand=True)
            if sname not in wb.sheet_names(): wb.release_resources(); return ''
            ws = wb.sheet_by_name(sname)
            primary = ''
            fallback = ''
            for r in range(min(ws.nrows, 8)):
                for c in range(ws.ncols):
                    v = ws.cell(r, c).value
                    if isinstance(v, str):
                        s = v.strip()
                        if not s or len(s) <= 20 or _FRENCH_RE.search(s):
                            continue
                        fw = s.split()[0].rstrip('.,;:!?')
                        if norm_qcode(fw) == sn_norm:
                            primary = s; break
                        if not fallback and not _HEADER_SKIP.match(s):
                            fallback = s
                if primary: break
            wb.release_resources()
            return primary or fallback
    except Exception:
        return ''

def get_all_sheet_questions(fpath, sheets):
    """Return {norm_qcode: (sname, [text, ...])} — open workbook ONCE for all sheets.

    Stores ALL candidate texts per sheet (multiple languages, fallbacks) so that
    the caller can try each one against the DB text.
    """
    result = {}
    skip = {'content', 'index', 'toc', 'contents'}
    _HEADER_SKIP = re.compile(
        r'^(eurobarometer|vol\s*a|base:|terrain|fieldwork|<<|ue\d|eu\d)',
        re.IGNORECASE
    )
    try:
        if fpath.lower().endswith('.xlsx'):
            if not _XLSX: return {}
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            for sname in sheets:
                if sname.lower() in skip or sname not in wb.sheetnames:
                    continue
                sn_norm = norm_qcode(sname)
                ws = wb[sname]
                candidates = []  # all long-enough non-header cells
                for row in ws.iter_rows(values_only=True, max_row=8):
                    for cell in row:
                        if cell and isinstance(cell, str):
                            s = cell.strip()
                            if not s or len(s) <= 20:
                                continue
                            if _HEADER_SKIP.match(s):
                                continue
                            candidates.append(s)
                result[sn_norm] = (sname, candidates)
            wb.close()
        else:
            if not _XLS: return {}
            wb = xlrd.open_workbook(fpath, on_demand=True)
            avail = set(wb.sheet_names())
            for sname in sheets:
                if sname.lower() in skip or sname not in avail:
                    continue
                sn_norm = norm_qcode(sname)
                ws = wb.sheet_by_name(sname)
                candidates = []
                for r in range(min(ws.nrows, 8)):
                    for c in range(ws.ncols):
                        v = ws.cell(r, c).value
                        if isinstance(v, str):
                            s = v.strip()
                            if not s or len(s) <= 20:
                                continue
                            if _HEADER_SKIP.match(s):
                                continue
                            candidates.append(s)
                result[sn_norm] = (sname, candidates)
            wb.release_resources()
    except Exception as e:
        print(f"  [WARN] batch read failed for {os.path.basename(fpath)}: {e}")
    return result

# ── load/save overrides ────────────────────────────────────────────────────────
def load_overrides():
    try:
        with open(OVERRIDES_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}

def save_overrides(ov):
    with open(OVERRIDES_FILE, 'w', encoding='utf-8') as f:
        json.dump(ov, f, indent=2)
    print(f"\nOverrides saved to {OVERRIDES_FILE}")

# ── check if a question already has a match via existing logic ─────────────────
def has_existing_match(wave_key, qnum, qtext, file_map, existing_ov):
    """Return True if this question is already resolved."""
    # Manual override exists
    if qnum in existing_ov.get(wave_key, {}):
        return True
    # Question index (pre-built cache on disk)
    try:
        with open(QN_INDEX_FILE, 'r', encoding='utf-8') as f:
            qidx = json.load(f)
        if qnum in qidx.get(wave_key, {}):
            return True
    except Exception:
        pass
    # Sheet name direct match (sheet name == qnum after normalisation)
    files = file_map.get(wave_key, [])
    qn = norm_qcode(qnum)
    for fpath in files:
        for sname in get_sheets(fpath):
            if norm_qcode(sname) == qn:
                xl = get_sheet_text(fpath, sname)
                if texts_match(qtext, xl):
                    return True
    return False

# ── main ───────────────────────────────────────────────────────────────────────
def main():
    # Build wave -> files map
    file_map = {}
    for fname in os.listdir(VOL_A_DIR):
        if not (fname.lower().endswith('.xls') or fname.lower().endswith('.xlsx')):
            continue
        w = wave_from_file(fname)
        if w:
            file_map.setdefault(w, []).append(os.path.join(VOL_A_DIR, fname))

    existing_ov = load_overrides()

    # Get all wave/question/text from DB
    con = duckdb.connect(DB_PATH, read_only=True)
    rows = con.execute(
        'SELECT DISTINCT "Wave","Question Number","Question(s)" FROM enes '
        'WHERE "Wave" IS NOT NULL AND "Question Number" IS NOT NULL'
    ).fetchall()
    con.close()

    from collections import defaultdict
    db = defaultdict(list)
    for wave, qnum, qtext in rows:
        key = norm_wave(str(wave).strip())
        db[key].append((str(qnum).strip(), str(qtext or '').strip()))

    new_overrides = {}   # wave_key -> {qnum: [fpath, sname]}
    stats = {'ok':0, 'fixed':0, 'unfixable':0, 'no_file':0}

    waves = sorted(k for k in db if (not WAVE_FILTER or WAVE_FILTER in k.lower()))

    for wave_key in waves:
        files = file_map.get(wave_key, [])
        if not files:
            stats['no_file'] += len(db[wave_key])
            continue

        # Load all sheets + texts for this wave once (open each file once)
        wave_sheets = {}  # fpath -> {norm_qcode: (sname, text)}
        for fpath in files:
            sheets = get_sheets(fpath)
            wave_sheets[fpath] = get_all_sheet_questions(fpath, sheets)

        ov_for_wave = existing_ov.get(wave_key, {})
        wave_new_ov = {}

        for qnum, qtext in db[wave_key]:
            # Skip already overridden
            if qnum in ov_for_wave:
                stats['ok'] += 1
                continue

            qn = norm_qcode(qnum)

            def any_match(qtext, candidates):
                return any(texts_match(qtext, t) for t in candidates)

            # Check if already matched by sheet name + text
            already = False
            for fpath, sq in wave_sheets.items():
                if qn in sq:
                    sname, xl_texts = sq[qn]
                    if any_match(qtext, xl_texts):
                        already = True
                        break
            if already:
                stats['ok'] += 1
                continue

            # --- UNMATCHED: try to fix ---
            fixed = False

            # Search all sheets by text
            best = None
            for fpath, sq in wave_sheets.items():
                for norm, (sname, xl_texts) in sq.items():
                    if any_match(qtext, xl_texts):
                        best = (fpath, sname)
                        break
                if best:
                    break

            if best:
                fpath, sname = best
                wave_new_ov[qnum] = [os.path.basename(fpath), sname]
                print(f"  [FIXED]  EB{wave_key} Q{qnum} -> {os.path.basename(fpath)} / {sname}")
                stats['fixed'] += 1
                fixed = True

            if not fixed:
                # List available questions in that wave's Excel
                available = []
                for fpath, sq in wave_sheets.items():
                    for norm, (sname, xl_texts) in sq.items():
                        first = xl_texts[0] if xl_texts else ''
                        available.append(f"{sname}: {first[:60]!r}")
                print(f"  [MISS]   EB{wave_key} Q{qnum} | DB: {qtext[:60]!r}")
                if available:
                    print(f"           Available in Excel ({len(available)} sheets):")
                    for a in available[:15]:
                        print(f"             {a}")
                    if len(available) > 15:
                        print(f"             ... ({len(available)-15} more)")
                else:
                    print(f"           (no readable sheets found)")
                stats['unfixable'] += 1

        if wave_new_ov:
            new_overrides[wave_key] = wave_new_ov

    # Merge new overrides into existing
    if new_overrides:
        for wave_key, qmap in new_overrides.items():
            existing_ov.setdefault(wave_key, {}).update(qmap)
        save_overrides(existing_ov)
        total_new = sum(len(v) for v in new_overrides.values())
        print(f"\nAdded {total_new} new overrides across {len(new_overrides)} waves.")
    else:
        print("\nNo new overrides needed.")

    print(f"\nSUMMARY")
    print(f"  Already matched : {stats['ok']}")
    print(f"  Newly fixed     : {stats['fixed']}")
    print(f"  Still unmatched : {stats['unfixable']}")
    print(f"  No Vol A file   : {stats['no_file']}")

if __name__ == '__main__':
    main()
