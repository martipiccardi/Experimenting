"""Volume A Excel file lookup and HTML rendering.

Supports both .xls (via xlrd) and .xlsx (via openpyxl).
Multiple files can share the same EB wave number — all are searched
for the requested question sheet.

Sheet names are cached at startup: {wave_key: {filepath: [sheet_names]}}
so all lookups are O(1) dict access, no file I/O per request.
"""
import gzip
import json
import os
import re

_GATE_VERSION = "global-gate-block-v1"  # bump to verify server has latest code
import threading

try:
    import xlrd
    _XLRD_OK = True
except ImportError:
    _XLRD_OK = False

try:
    import openpyxl
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

_DEFAULT_VOL_A_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "..", "..", "data", "vol_a"
)
VOL_A_DIR = os.environ.get("VOL_A_DIR", _DEFAULT_VOL_A_DIR)

# Manual match overrides: data/vol_a_overrides.json
# Format: { "eb97.2": { "QB1": ["filename.xlsx", "SheetName"] }, ... }
# Use [null, null] to force TOC (no Vol A sheet for this question).
_OVERRIDES_FILE = os.path.join(os.path.dirname(VOL_A_DIR), "vol_a_overrides.json")
_overrides: dict = {}  # {normalized_wave: {question: [fname, sheet] | None}}


def _load_overrides() -> None:
    global _overrides
    try:
        with open(_OVERRIDES_FILE, 'r', encoding='utf-8') as f:
            raw = json.load(f)
        result = {}
        for wave, qmap in raw.items():
            if wave.startswith('_'):
                continue  # skip comment keys
            wkey = _normalize_wave(wave)  # same as render_sheet_as_html uses
            result[wkey] = {}
            for qn, entry in qmap.items():
                if qn.startswith('_'):
                    continue  # skip comment keys inside wave dict
                if entry is None or (isinstance(entry, list) and not any(entry)):
                    result[wkey][qn] = None  # → TOC
                else:
                    fname, sheet = entry[0], entry[1]
                    result[wkey][qn] = [os.path.join(VOL_A_DIR, fname), sheet]
        _overrides = result
        print(f"[vol_a] Overrides loaded: {sum(len(v) for v in _overrides.values())} entries", flush=True)
    except FileNotFoundError:
        _overrides = {}
    except Exception as e:
        print(f"[vol_a] Failed to load overrides: {e}", flush=True)
        _overrides = {}


# Disk cache for sheet map — survives restarts, lives next to vol_a/ dir
_SHEET_MAP_CACHE_FILE = os.path.join(os.path.dirname(VOL_A_DIR), "vol_a_sheet_map.json")
_QUESTION_INDEX_CACHE_FILE = os.path.join(os.path.dirname(VOL_A_DIR), "vol_a_question_index.json")
_TEXT_INDEX_CACHE_FILE = os.path.join(os.path.dirname(VOL_A_DIR), "vol_a_text_index.json")

# Disk HTML cache — gzipped rendered HTML, survives restarts AND deployments.
# On Azure: /home/vol_a_html_cache (set via env var in startup.sh, outside wwwroot).
# Locally: data/vol_a_html_cache (next to vol_a/).
_HTML_CACHE_DIR = os.environ.get(
    "VOL_A_HTML_CACHE_DIR",
    os.path.join(os.path.dirname(VOL_A_DIR), "vol_a_html_cache"),
)

# Bump this string whenever the HTML rendering changes (chart buttons, layout, etc.).
# On startup, if the cache version file doesn't match, all cached HTML is wiped
# so pages are re-rendered with the new code.
_HTML_CACHE_VERSION = "v17-eu28-fix-2026"

def _check_html_cache_version():
    """Wipe disk HTML cache if the stored version doesn't match _HTML_CACHE_VERSION."""
    if not os.path.isdir(_HTML_CACHE_DIR):
        return
    version_file = os.path.join(_HTML_CACHE_DIR, ".version")
    try:
        stored = open(version_file).read().strip() if os.path.exists(version_file) else ""
    except Exception:
        stored = ""
    if stored == _HTML_CACHE_VERSION:
        return
    print(f"[vol_a] HTML cache version mismatch ({stored!r} != {_HTML_CACHE_VERSION!r}) — clearing cache.", flush=True)
    import glob
    for f in glob.glob(os.path.join(_HTML_CACHE_DIR, "**", "*.gz"), recursive=True):
        try:
            os.remove(f)
        except Exception:
            pass
    try:
        with open(version_file, "w") as fh:
            fh.write(_HTML_CACHE_VERSION)
        print("[vol_a] HTML cache cleared and version file updated.", flush=True)
    except Exception as e:
        print(f"[vol_a] Could not write version file: {e}", flush=True)

# In-memory cache: {wave_key: {filepath: [sheet_name, ...]}}
_wave_sheet_map = None
_sheet_map_lock = threading.Lock()

# Supplementary index: {wave_key: {question_number: [filepath, sheet_name]}}
# Built from Index sheets for old-format files where sheet names != question numbers.
_question_index = None

# Text-based index: {wave_key: {text_fingerprint: [filepath, sheet_name]}}
# Built from Content sheet col2 (English question text) for T-format files.
_text_index = None

# Lazy cache: {(fpath, sheet_name): text_fingerprint or ''}
# Populated on-demand when letter-suffix matches are verified.
# Avoids the long upfront build time of a full sheet scan.
_sheet_eng_cache: dict = {}

# Cache: {(wave_key, question) -> rendered HTML string}
_html_cache = {}

# Cache: {(fpath, sheet_name) -> (label, normalized_text) | None}
# label = question code extracted from the Excel cell (e.g. 'QB1', 'QA3')
# normalized_text = whitespace-collapsed, lowercase, prefix-stripped question text
_match_cache: dict = {}

# French-specific accented characters — used to distinguish French from English cells
_FRENCH_CHARS_RE = re.compile(r'[éèêëàâäïîôùûüçœæÉÈÊËÀÂÄÏÎÔÙÛÜÇŒÆ]')

# Regex to extract a question number from the start of an Index cell
# e.g. "Q48 Tous les combien..." → "Q48"
#      "D1.1 A propos de..." → "D1.1"
_QN_RE = re.compile(r'^([A-Za-z]{1,4}\d+[a-zA-Z]?(?:[._]\d+)?)\s')

# Regex to capture the question-number prefix from a question cell, e.g.:
#   "QB1. How satisfied..." → "QB1"
#   "QC2b.1. Do you think..." → "QC2b.1"
_PREFIX_RE = re.compile(
    r'^([A-Za-z]{1,4}\d+[a-zA-Z0-9]*(?:[._]\d+[a-zA-Z0-9]*)*)\s*[.:]?\s+'
)


# ── Wave number helpers ────────────────────────────────────────────────────────

def _normalize_wave(wave: str) -> str:
    """'EB 100.2' or 'EB100.2' → '100.2'"""
    w = wave.strip()
    w = re.sub(r'^[Ee][Bb]\s*', '', w).strip()
    return w


def _wave_from_filename(fname: str):
    """Extract EB wave number from filename.

    Old format: 'EB62.1_EBS215_Lisbon_VOL_A.xls'  → '62.1'
    New format: 'eb1001_Combined_vol_A.xlsx'        → '100.1'
                'eb981_vol_A.xlsx'                  → '98.1'
    """
    # Old format: explicit EB + dotted number
    m = re.search(r'[Ee][Bb](\d+\.\d+)', fname)
    if m:
        return m.group(1)
    # New format: eb{major}{minor} e.g. eb1001 → 100.1, eb981 → 98.1
    m = re.match(r'eb(\d{3,4})_', fname, re.IGNORECASE)
    if m:
        digits = m.group(1)
        major, minor = digits[:-1], digits[-1]
        return f"{major}.{minor}"
    return None


def _wave_from_content_sheet(filepath: str):
    """Fallback for .xlsx files without EB number in filename: read Content sheet."""
    if not _OPENPYXL_OK:
        return None
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet_name = next(
            (s for s in ('Content', 'content', 'CONTENT') if s in wb.sheetnames),
            wb.sheetnames[0] if wb.sheetnames else None
        )
        if sheet_name:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                if row and str(row[0]).strip().lower() == 'wave:':
                    val = row[1] if len(row) > 1 else None
                    if val is not None:
                        wb.close()
                        return str(val).strip()
        eb_pattern = re.compile(r'[Ee]urobarometer\s+(\d+\.\d+)')
        for sname in wb.sheetnames:
            ws = wb[sname]
            for row in ws.iter_rows(values_only=True, max_row=3):
                for cell in row:
                    if cell:
                        m = eb_pattern.search(str(cell))
                        if m:
                            wb.close()
                            return m.group(1)
        wb.close()
    except Exception:
        pass
    return None


# ── Sheet-name loading (fast, no cell data) ───────────────────────────────────

def _load_sheet_names_only(fpath: str):
    """Open file header only and return list of sheet names."""
    ext = fpath.lower().rsplit('.', 1)[-1]
    if ext == 'xlsx':
        if not _OPENPYXL_OK:
            return []
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            names = list(wb.sheetnames)
            wb.close()
            return names
        except Exception:
            return []
    else:
        if not _XLRD_OK:
            return []
        try:
            # on_demand=True: loads workbook structure only, skips cell data — fast
            wb = xlrd.open_workbook(fpath, on_demand=True)
            names = wb.sheet_names()
            wb.release_resources()
            return names
        except Exception:
            return []


# ── Sheet-name cache (built once at startup) ──────────────────────────────────

def _build_wave_sheet_map():
    """Scan VOL_A_DIR, open every Volume-A file once to cache sheet names.

    Returns {wave_key: {filepath: [sheet_names]}}.
    """
    result = {}
    if not os.path.isdir(VOL_A_DIR):
        return result

    for fname in os.listdir(VOL_A_DIR):
        ext = fname.lower().rsplit('.', 1)[-1]
        if ext not in ('xls', 'xlsx'):
            continue
        # Only Volume A: name must contain vol_A or volume_A (not VOL_AA, VOL_B, etc.)
        # Matches: VOL_A, vol_A, volA, "vol A", volume_A, Volume_A — excludes: VOL_AA, VOL_B, etc.
        if not re.search(r'vol(?:ume)?[_ ]?a(?!a)', fname, re.IGNORECASE):
            continue

        fpath = os.path.join(VOL_A_DIR, fname)
        wave_key = _wave_from_filename(fname)
        if not wave_key and ext == 'xlsx':
            wave_key = _wave_from_content_sheet(fpath)
        if not wave_key:
            continue

        sheets = _load_sheet_names_only(fpath)
        result.setdefault(wave_key, {})[fpath] = sheets

    return result


def _save_sheet_map(m):
    try:
        with open(_SHEET_MAP_CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(m, f)
        print(f"[vol_a] Sheet map saved to disk ({len(m)} waves)", flush=True)
    except Exception as e:
        print(f"[vol_a] Warning: could not save sheet map cache: {e}", flush=True)


def _load_sheet_map_from_disk():
    try:
        with open(_SHEET_MAP_CACHE_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        # Re-resolve any stale absolute paths (e.g. Windows paths on Linux).
        # Extract just the filename and rebuild the path using the current VOL_A_DIR.
        fixed = {}
        needs_fix = False
        for wave, file_dict in data.items():
            fixed[wave] = {}
            for fpath, sheets in file_dict.items():
                if os.path.exists(fpath):
                    fixed[wave][fpath] = sheets
                else:
                    fname = os.path.basename(fpath.replace('\\', '/'))
                    resolved = os.path.join(VOL_A_DIR, fname)
                    fixed[wave][resolved] = sheets
                    needs_fix = True
        if needs_fix:
            print("[vol_a] Sheet map paths re-resolved for current platform", flush=True)
        print(f"[vol_a] Sheet map loaded from disk cache: {len(fixed)} waves", flush=True)
        return fixed
    except Exception:
        return None


# ── Question index (Index-sheet parsing for old-format files) ─────────────────

def _extract_qn_from_index(fpath: str, sheet_names: list) -> dict:
    """Parse the Index sheet of a Vol A file to extract {question_number: sheet_name}.

    Works for old-format files where sheet names (S2, S3…) don't match question
    numbers (Q48, D1…). In that case every question maps to the first data sheet.

    Skips newer-format files where the Index first column contains sheet references
    like "QB1'!A1" — those are already handled by exact/prefix matching.

    Returns {} if nothing useful can be extracted.
    """
    if 'Index' not in sheet_names:
        return {}

    # Data sheets: skip meta sheets
    _META = {'index', 'bookmarks', 'content', 'b', 'b2', 'toc'}
    data_sheets = [s for s in sheet_names if s.lower() not in _META]
    if not data_sheets:
        return {}

    ext = fpath.lower().rsplit('.', 1)[-1]
    mapping = {}

    try:
        if ext == 'xlsx':
            if not _OPENPYXL_OK:
                return {}
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            if 'Index' not in wb.sheetnames:
                wb.close()
                return {}
            ws = wb['Index']
            for row in ws.iter_rows(values_only=True):
                if not row or not row[0]:
                    continue
                cell = str(row[0]).strip()
                # Skip newer-format sheet references ("QB1'!A1")
                if re.match(r"'?[^'!]+'?!A\d+$", cell):
                    continue
                m = _QN_RE.match(cell)
                if m:
                    qn = m.group(1)
                    # Prefer a data sheet whose name matches the question number
                    target = next(
                        (s for s in data_sheets
                         if s == qn or s.startswith(qn + '.') or s.startswith(qn + '_')),
                        data_sheets[0]
                    )
                    mapping[qn] = target
            wb.close()
        else:
            if not _XLRD_OK:
                return {}
            wb = xlrd.open_workbook(fpath, on_demand=True)
            if 'Index' not in wb.sheet_names():
                wb.release_resources()
                return {}
            ws = wb.sheet_by_name('Index')
            for r in range(ws.nrows):
                cell = str(ws.cell(r, 0).value).strip()
                if not cell:
                    continue
                if re.match(r"'?[^'!]+'?!A\d+$", cell):
                    continue
                m = _QN_RE.match(cell)
                if m:
                    qn = m.group(1)
                    target = next(
                        (s for s in data_sheets
                         if s == qn or s.startswith(qn + '.') or s.startswith(qn + '_')),
                        data_sheets[0]
                    )
                    mapping[qn] = target
            wb.release_resources()
    except Exception:
        pass

    return mapping


def _extract_qn_from_content(fpath: str, sheet_names: list) -> dict:
    """Parse the Content sheet of a Vol A file to extract {question_number: sheet_name}.

    Works for files (e.g. EB93-94) where data sheets are named T1, T2, T3… and
    the Content sheet maps each T-sheet to the real question number:
        Sheet | Label
        T3    | QC1 Regarding smoking...
        T4    | QC2a Do you...

    Returns {} if the file has no T-format sheets or no Content sheet.
    """
    # Only process files that have T-format sheet names (T1, T2, …)
    t_sheets = {s for s in sheet_names if re.match(r'^T\d+$', s)}
    if not t_sheets:
        return {}

    content_sheet = next(
        (s for s in sheet_names if s.lower() == 'content'),
        None
    )
    if not content_sheet:
        return {}

    ext = fpath.lower().rsplit('.', 1)[-1]
    mapping = {}   # {question_number: t_sheet_name}

    try:
        if ext == 'xlsx':
            if not _OPENPYXL_OK:
                return {}
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            if content_sheet not in wb.sheetnames:
                wb.close()
                return {}
            ws = wb[content_sheet]
            header_found = False
            for row in ws.iter_rows(values_only=True):
                if not row:
                    continue
                col0 = str(row[0]).strip() if row[0] is not None else ''
                if not header_found:
                    if col0.lower() == 'sheet':
                        header_found = True
                    continue
                # Data row: col0 = T-sheet name, col1+ = question label
                if col0 not in t_sheets:
                    continue
                for ci in range(1, min(5, len(row))):
                    if row[ci] is None:
                        continue
                    label = str(row[ci]).strip()
                    m = _QN_RE.match(label)
                    if m:
                        qn = m.group(1)
                        if qn != col0:   # skip if label accidentally equals sheet name
                            mapping[qn] = col0
                        break
            wb.close()
        else:
            if not _XLRD_OK:
                return {}
            wb = xlrd.open_workbook(fpath, on_demand=True)
            if content_sheet not in wb.sheet_names():
                wb.release_resources()
                return {}
            ws = wb.sheet_by_name(content_sheet)
            header_found = False
            for r in range(ws.nrows):
                col0 = str(ws.cell(r, 0).value).strip()
                if not header_found:
                    if col0.lower() == 'sheet':
                        header_found = True
                    continue
                if col0 not in t_sheets:
                    continue
                for ci in range(1, min(5, ws.ncols)):
                    label = str(ws.cell(r, ci).value).strip()
                    m = _QN_RE.match(label)
                    if m:
                        qn = m.group(1)
                        if qn != col0:
                            mapping[qn] = col0
                        break
            wb.release_resources()
    except Exception:
        pass

    return mapping


def _extract_qn_from_sheet_cells(fpath: str, sheet_names: list) -> dict:
    """Scan the first rows of every data sheet for question numbers in cell content.

    Complements the Content/Index strategies: finds question numbers that appear
    inside the actual table header cells (e.g. 'QB8.1 Some question text…' in
    row 3 col 2 of sheet T16) even when that mapping is absent from the Content
    sheet.

    Returns {question_number: sheet_name}.
    """
    _SKIP = {
        'content', 'index', 'toc', 'title', 'notes', 'note', 'cover',
        'annex', 'annexe', 'légende', 'legende', 'legend', 'intro',
    }
    ext = fpath.lower().rsplit('.', 1)[-1]
    mapping = {}

    try:
        if ext == 'xlsx':
            if not _OPENPYXL_OK:
                return {}
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                if sheet_name.lower() in _SKIP:
                    continue
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True, max_row=6):
                    for cell in row[:6]:
                        if cell is None:
                            continue
                        m = _QN_RE.match(str(cell).strip())
                        if m:
                            qn = m.group(1)
                            # Don't self-reference (sheet QB1 finding QB1 in its own cell)
                            if qn != sheet_name and qn not in mapping:
                                mapping[qn] = sheet_name
            wb.close()
        else:
            if not _XLRD_OK:
                return {}
            wb = xlrd.open_workbook(fpath, on_demand=True)
            for sheet_name in wb.sheet_names():
                if sheet_name.lower() in _SKIP:
                    continue
                ws = wb.sheet_by_name(sheet_name)
                nrows = min(ws.nrows, 6)
                ncols = min(ws.ncols, 6)
                for r in range(nrows):
                    for c in range(ncols):
                        text = str(ws.cell(r, c).value).strip()
                        m = _QN_RE.match(text)
                        if m:
                            qn = m.group(1)
                            if qn != sheet_name and qn not in mapping:
                                mapping[qn] = sheet_name
            wb.release_resources()
    except Exception:
        pass

    return mapping


# ── Text-based index (English question text → T-sheet) ────────────────────────

def _text_fingerprint(text: str, n: int = 70) -> str:
    """Normalize question text into a compact fingerprint for fuzzy matching.

    Strips instruction lines (everything after the first blank line), strips
    the question-number prefix (e.g. 'QA1 ', 'QB8.1 ', 'QC2b.1. '), lowercases,
    collapses whitespace, and takes the first n characters.
    """
    # Strip instruction text: take only the first paragraph (before any \n\n)
    text = text.split('\n\n')[0].strip()
    # Strip leading question number including variants like QC2b.1. and QE3bT.
    # Pattern: 1-4 letters, digits, optional alphanumeric suffix, optional
    # repeated .N or _N groups, optional trailing punctuation + whitespace.
    text = re.sub(
        r'^[A-Za-z]{1,4}\d+[a-zA-Z0-9]*(?:[._]\d+[a-zA-Z0-9]*)*\s*[.:]?\s*',
        '', text,
    )
    text = re.sub(r'\s+', ' ', text.lower()).strip()
    return text[:n]


def _normalize_exact(text: str) -> str:
    """Normalize question text for exact matching: strip instruction blocks that
    follow a blank line (\\n\\n), strip QN prefix, collapse all whitespace to
    single spaces, lowercase.  No truncation."""
    # Take only the part before the first blank-line separator so that
    # DB texts like "How satisfied…?\n\n(SHOW SCREEN - READ OUT)" become just
    # "How satisfied…?" (matching the shorter Excel cell text).
    text = text.split('\n\n')[0].strip()
    text = re.sub(
        r'^[A-Za-z]{1,4}\d+[a-zA-Z0-9]*(?:[._]\d+[a-zA-Z0-9]*)*\s*[.:]?\s*',
        '', text,
    )
    return re.sub(r'\s+', ' ', text.lower()).strip()


def _text_similarity(a: str, b: str) -> float:
    """Rough token-overlap similarity between two text strings."""
    if not a or not b:
        return 0.0
    a_tokens = set(re.findall(r'\w+', a.lower()))
    b_tokens = set(re.findall(r'\w+', b.lower()))
    if not a_tokens or not b_tokens:
        return 0.0
    overlap = len(a_tokens & b_tokens)
    return overlap / max(len(a_tokens), len(b_tokens))


def _normalize_qcode(s: str) -> str:
    """Normalize a question code for comparison: remove dots/underscores, lowercase."""
    return re.sub(r'[._]', '', s).lower()


def _extract_english_question_text(fpath: str, sheet_name: str) -> str | None:
    """Find the English question text from the header area of a data sheet.

    Scans the first 6 rows for a cell that:
    - Starts with the sheet name (allowing dot vs underscore normalization)
    - Contains no French-specific accented characters (é, è, à, ç, …)
    - Is longer than 20 characters

    Returns the English question text string, or None if not found.
    """
    sn_norm = _normalize_qcode(sheet_name)
    ext = fpath.lower().rsplit('.', 1)[-1]

    def _is_english_question_cell(s: str) -> bool:
        if not s or len(s) <= 20:
            return False
        if _FRENCH_CHARS_RE.search(s):
            return False
        first_word = s.split()[0].rstrip('.,;:!?') if s.split() else ''
        return _normalize_qcode(first_word) == sn_norm

    try:
        if ext == 'xlsx':
            if not _OPENPYXL_OK:
                return None
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return None
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True, max_row=6):
                for cell in row:
                    if cell and isinstance(cell, str):
                        s = cell.strip()
                        if _is_english_question_cell(s):
                            wb.close()
                            return s
            wb.close()
        else:
            if not _XLRD_OK:
                return None
            wb = xlrd.open_workbook(fpath, on_demand=True)
            if sheet_name not in wb.sheet_names():
                wb.release_resources()
                return None
            ws = wb.sheet_by_name(sheet_name)
            for r in range(min(ws.nrows, 6)):
                for c in range(ws.ncols):
                    v = ws.cell(r, c).value
                    if isinstance(v, str):
                        s = v.strip()
                        if _is_english_question_cell(s):
                            wb.release_resources()
                            return s
            wb.release_resources()
    except Exception:
        pass
    return None


def _batch_load_match_for_file(fpath: str, sheet_names: list) -> None:
    """Open fpath once and fill _match_cache for all requested sheets.

    For each sheet, scans the first 6 rows for a cell that:
    - is longer than 20 characters
    - contains no French accented characters
    Extracts the question-number label (e.g. 'QB1') and the normalized question
    text (prefix stripped, whitespace collapsed, lowercase).
    Stores (label, normalized_text) or None in _match_cache.
    """
    ext = fpath.lower().rsplit('.', 1)[-1]

    def _process(s: str):
        if not s or len(s) <= 20 or _FRENCH_CHARS_RE.search(s):
            return None
        # Must start with a QN prefix (QB1a., QA1., D1., T3. …)
        # This skips fieldwork date rows, wave info rows, etc.
        m = _PREFIX_RE.match(s)
        if not m:
            return None
        label = m.group(1)
        normalized = _normalize_exact(s)
        return (label, normalized) if normalized else None

    try:
        if ext == 'xlsx':
            if not _OPENPYXL_OK:
                for sn in sheet_names:
                    _match_cache[(fpath, sn)] = None
                return
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            for sn in sheet_names:
                result = None
                if sn in wb.sheetnames:
                    ws = wb[sn]
                    for row in ws.iter_rows(values_only=True, max_row=6):
                        for cell in row:
                            if cell and isinstance(cell, str):
                                result = _process(cell.strip())
                                if result:
                                    break
                        if result:
                            break
                _match_cache[(fpath, sn)] = result
            wb.close()
        else:
            if not _XLRD_OK:
                for sn in sheet_names:
                    _match_cache[(fpath, sn)] = None
                return
            wb = xlrd.open_workbook(fpath, on_demand=True)
            for sn in sheet_names:
                result = None
                if sn in wb.sheet_names():
                    ws = wb.sheet_by_name(sn)
                    for r in range(min(ws.nrows, 6)):
                        for c in range(ws.ncols):
                            v = ws.cell(r, c).value
                            if isinstance(v, str):
                                result = _process(v.strip())
                                if result:
                                    break
                        if result:
                            break
                _match_cache[(fpath, sn)] = result
            wb.release_resources()
    except Exception:
        for sn in sheet_names:
            if (fpath, sn) not in _match_cache:
                _match_cache[(fpath, sn)] = None


def _get_sheet_eng_fingerprint(fpath: str, sheet_name: str) -> str:
    """Return the English question text fingerprint for a sheet, with lazy caching.

    On first call for (fpath, sheet_name), opens the file and extracts the English
    question text from the header rows.  The result (fingerprint or '' if not found)
    is cached in _sheet_eng_cache so subsequent calls are O(1).
    """
    cache_key = (fpath, sheet_name)
    if cache_key in _sheet_eng_cache:
        return _sheet_eng_cache[cache_key]
    eng_text = _extract_english_question_text(fpath, sheet_name)
    if eng_text:
        fp = _text_fingerprint(eng_text)
        result = fp if len(fp) >= 15 else ''
    else:
        result = ''
    _sheet_eng_cache[cache_key] = result
    return result


def _batch_load_sheet_fps(fpath: str, sheet_names: list) -> None:
    """Open fpath once and populate _sheet_eng_cache for all listed sheets.

    Much faster than calling _get_sheet_eng_fingerprint separately for each sheet
    when many sheets from the same file need verification in one pass.
    """
    sheets_needed = [s for s in sheet_names if (fpath, s) not in _sheet_eng_cache]
    if not sheets_needed:
        return
    ext = fpath.lower().rsplit('.', 1)[-1]
    try:
        if ext == 'xlsx':
            if not _OPENPYXL_OK:
                for s in sheets_needed:
                    _sheet_eng_cache[(fpath, s)] = ''
                return
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            for sname in sheets_needed:
                if sname not in wb.sheetnames:
                    _sheet_eng_cache[(fpath, sname)] = ''
                    continue
                sn_norm = _normalize_qcode(sname)
                ws = wb[sname]
                found = ''
                for row in ws.iter_rows(values_only=True, max_row=6):
                    for cell in row:
                        if cell and isinstance(cell, str):
                            s = cell.strip()
                            if not s or len(s) <= 20 or _FRENCH_CHARS_RE.search(s):
                                continue
                            fw = s.split()[0].rstrip('.,;:!?') if s.split() else ''
                            if _normalize_qcode(fw) == sn_norm:
                                fp = _text_fingerprint(s)
                                found = fp if len(fp) >= 15 else ''
                                break
                    if found:
                        break
                _sheet_eng_cache[(fpath, sname)] = found
            wb.close()
        else:
            if not _XLRD_OK:
                for s in sheets_needed:
                    _sheet_eng_cache[(fpath, s)] = ''
                return
            wb = xlrd.open_workbook(fpath, on_demand=True)
            available = set(wb.sheet_names())
            for sname in sheets_needed:
                if sname not in available:
                    _sheet_eng_cache[(fpath, sname)] = ''
                    continue
                sn_norm = _normalize_qcode(sname)
                ws = wb.sheet_by_name(sname)
                found = ''
                for r in range(min(ws.nrows, 6)):
                    for c in range(ws.ncols):
                        v = ws.cell(r, c).value
                        if isinstance(v, str):
                            s = v.strip()
                            if not s or len(s) <= 20 or _FRENCH_CHARS_RE.search(s):
                                continue
                            fw = s.split()[0].rstrip('.,;:!?') if s.split() else ''
                            if _normalize_qcode(fw) == sn_norm:
                                fp = _text_fingerprint(s)
                                found = fp if len(fp) >= 15 else ''
                                break
                    if found:
                        break
                _sheet_eng_cache[(fpath, sname)] = found
            wb.release_resources()
    except Exception:
        for sname in sheets_needed:
            if (fpath, sname) not in _sheet_eng_cache:
                _sheet_eng_cache[(fpath, sname)] = ''


def _build_text_index(wave_sheet_map: dict) -> dict:
    """Build {wave_key: {text_fingerprint: [filepath, t_sheet]}} from Content sheet col2.

    Only processes T-format xlsx files that have a Content sheet with an English
    text column (col2, 1-indexed = row[2] 0-indexed).
    """
    result = {}
    for wave_key, file_sheets in wave_sheet_map.items():
        wave_map = {}
        for fpath, sheets in file_sheets.items():
            t_sheets = {s for s in sheets if re.match(r'^T\d+$', s)}
            if not t_sheets:
                continue
            content_sheet = next((s for s in sheets if s.lower() == 'content'), None)
            if not content_sheet or not fpath.lower().endswith('.xlsx'):
                continue
            if not _OPENPYXL_OK:
                continue
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                if content_sheet not in wb.sheetnames:
                    wb.close()
                    continue
                ws = wb[content_sheet]
                header_found = False
                for row in ws.iter_rows(values_only=True):
                    if not row:
                        continue
                    col0 = str(row[0]).strip() if row[0] is not None else ''
                    if not header_found:
                        if col0.lower() == 'sheet':
                            header_found = True
                        continue
                    if col0 not in t_sheets:
                        continue
                    # col2 (0-indexed) = English text column
                    if len(row) > 2 and row[2] is not None:
                        english_text = str(row[2]).strip()
                        fp = _text_fingerprint(english_text)
                        # Require at least 20 chars to avoid accidental matches
                        if len(fp) >= 20 and fp not in wave_map:
                            wave_map[fp] = [fpath, col0]
                wb.close()
            except Exception:
                pass
        if wave_map:
            result[wave_key] = wave_map
    return result


def _save_text_index(idx: dict):
    try:
        with open(_TEXT_INDEX_CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(idx, f)
        total = sum(len(v) for v in idx.values())
        print(f"[vol_a] Text index saved: {total} entries across {len(idx)} waves", flush=True)
    except Exception as e:
        print(f"[vol_a] Warning: could not save text index: {e}", flush=True)


def _load_text_index_from_disk():
    try:
        with open(_TEXT_INDEX_CACHE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return None


def get_text_index():
    global _text_index
    if _text_index is not None:
        return _text_index
    cached = _load_text_index_from_disk()
    if cached is not None:
        _text_index = cached
        return _text_index
    _text_index = _build_text_index(get_wave_sheet_map())
    _save_text_index(_text_index)
    return _text_index


# Module-level cache for DB question texts — avoids repeated DB connections
# when the gate is applied to many questions in the same process.
_db_text_cache: dict = {}


def _get_db_question_text(wave: str, question: str):
    """Look up the English question text from the DB for a given wave/question number."""
    cache_key = (wave, question)
    if cache_key in _db_text_cache:
        return _db_text_cache[cache_key]
    try:
        from data_store import get_conn, ensure_table as _ensure
        con = get_conn()
        _ensure(con)
        row = con.execute(
            'SELECT "Question(s)" FROM enes WHERE "Wave" = ? AND "Question Number" = ? LIMIT 1',
            [wave, question]
        ).fetchone()
        con.close()
        result = str(row[0]).strip() if row and row[0] else None
    except Exception:
        result = None
    _db_text_cache[cache_key] = result
    return result


def _build_question_index(wave_sheet_map: dict) -> dict:
    """Build {wave_key: {question_number: [filepath, sheet_name]}} from Index/Content sheets.

    Three strategies (in priority order):
      1. Index sheet     — old-format files (S2, S3…) where sheet names ≠ question numbers.
      2. Content sheet   — T-format files (EB93-94) where sheets are T1,T2… and the
                           Content sheet maps each T-sheet to the real question number.
      3. Cell scan       — fallback: scan first 6 rows × 6 cols of every data sheet for a
                           question-number pattern; catches questions missing from the
                           Content sheet or present in non-standard header layouts.

    Only adds entries not already resolvable via exact/prefix sheet-name matching.
    """
    result = {}
    for wave_key, file_sheets in wave_sheet_map.items():
        wave_map = {}
        for fpath, sheets in file_sheets.items():
            # --- Strategy 1: Index sheet (old S2/S3 format) ---
            qn_map = _extract_qn_from_index(fpath, sheets)
            for qn, sheet in qn_map.items():
                already_covered = (
                    qn in sheets or
                    any(s.startswith(qn + '.') or s.startswith(qn + '_') for s in sheets)
                )
                if not already_covered:
                    wave_map[qn] = [fpath, sheet]

            # --- Strategy 2: Content sheet (T-format EB93-94 style) ---
            content_map = _extract_qn_from_content(fpath, sheets)
            for qn, t_sheet in content_map.items():
                already_covered = (
                    qn in sheets or
                    any(s.startswith(qn + '.') or s.startswith(qn + '_') for s in sheets)
                )
                if not already_covered and qn not in wave_map:
                    wave_map[qn] = [fpath, t_sheet]

            # --- Strategy 3: Cell scan — catch questions missing from Content/Index ---
            cell_map = _extract_qn_from_sheet_cells(fpath, sheets)
            for qn, sheet in cell_map.items():
                already_covered = (
                    qn in sheets or
                    any(s.startswith(qn + '.') or s.startswith(qn + '_') for s in sheets)
                )
                if not already_covered and qn not in wave_map:
                    wave_map[qn] = [fpath, sheet]

        if wave_map:
            result[wave_key] = wave_map
    return result


def _save_question_index(idx: dict):
    try:
        with open(_QUESTION_INDEX_CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(idx, f)
        total = sum(len(v) for v in idx.values())
        print(f"[vol_a] Question index saved: {total} entries across {len(idx)} waves", flush=True)
    except Exception as e:
        print(f"[vol_a] Warning: could not save question index: {e}", flush=True)


def _load_question_index_from_disk():
    try:
        with open(_QUESTION_INDEX_CACHE_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        # Re-resolve stale absolute paths (e.g. Windows paths on Linux).
        fixed = {}
        for wave, qmap in data.items():
            fixed[wave] = {}
            for qn, entry in qmap.items():
                fpath, sheet = entry[0], entry[1]
                if not os.path.exists(fpath):
                    fname = os.path.basename(fpath.replace('\\', '/'))
                    fpath = os.path.join(VOL_A_DIR, fname)
                fixed[wave][qn] = [fpath, sheet]
        return fixed
    except Exception:
        return None


def get_question_index():
    global _question_index
    if _question_index is not None:
        return _question_index
    cached = _load_question_index_from_disk()
    if cached is not None:
        _question_index = cached
        return _question_index
    # Build from wave sheet map (which must already be loaded)
    _question_index = _build_question_index(get_wave_sheet_map())
    _save_question_index(_question_index)
    return _question_index


def get_wave_sheet_map():
    global _wave_sheet_map
    if _wave_sheet_map is not None:
        return _wave_sheet_map
    with _sheet_map_lock:
        if _wave_sheet_map is not None:
            return _wave_sheet_map
        _check_html_cache_version()  # wipe stale cached HTML if version changed
        _load_overrides()  # load once at startup alongside the sheet map
        cached = _load_sheet_map_from_disk()
        if cached is not None:
            _wave_sheet_map = cached
        else:
            print("[vol_a] No disk cache — scanning files (this may take a few minutes)…", flush=True)
            _wave_sheet_map = _build_wave_sheet_map()
            _save_sheet_map(_wave_sheet_map)
    return _wave_sheet_map


def reload_wave_file_map():
    """Force-rebuild the cache (call after uploading new files via Kudu)."""
    global _wave_sheet_map, _question_index, _text_index, _sheet_eng_cache, _html_cache, _match_cache
    _load_overrides()
    _wave_sheet_map = _build_wave_sheet_map()
    _save_sheet_map(_wave_sheet_map)
    _question_index = _build_question_index(_wave_sheet_map)
    _save_question_index(_question_index)
    _text_index = _build_text_index(_wave_sheet_map)
    _save_text_index(_text_index)
    _sheet_eng_cache = {}
    _html_cache = {}
    _match_cache = {}
    _clear_html_disk_cache()
    return _wave_sheet_map


def _html_disk_path(wave_key: str, sheet_name: str) -> str:
    """Return the gzip file path for a cached HTML snippet."""
    safe_wave = re.sub(r'[^A-Za-z0-9_-]', '_', wave_key)
    safe_sheet = re.sub(r'[^A-Za-z0-9_-]', '_', sheet_name)
    wave_dir = os.path.join(_HTML_CACHE_DIR, safe_wave)
    os.makedirs(wave_dir, exist_ok=True)
    return os.path.join(wave_dir, safe_sheet + '.html.gz')


def _load_html_from_disk(wave_key: str, sheet_name: str):
    """Load cached HTML from disk; return None if not found."""
    try:
        path = _html_disk_path(wave_key, sheet_name)
        if os.path.exists(path):
            with gzip.open(path, 'rt', encoding='utf-8') as f:
                return f.read()
    except Exception:
        pass
    return None


def _save_html_to_disk(wave_key: str, sheet_name: str, html: str) -> None:
    """Save rendered HTML to disk cache (gzipped, ~4-8× smaller)."""
    try:
        path = _html_disk_path(wave_key, sheet_name)
        with gzip.open(path, 'wt', encoding='utf-8') as f:
            f.write(html)
    except Exception:
        pass


def _clear_html_disk_cache() -> None:
    """Delete all gzipped HTML files from the disk cache directory."""
    import shutil
    try:
        if os.path.exists(_HTML_CACHE_DIR):
            shutil.rmtree(_HTML_CACHE_DIR)
            print(f"[vol_a] HTML disk cache cleared: {_HTML_CACHE_DIR}", flush=True)
    except Exception as e:
        print(f"[vol_a] Failed to clear HTML disk cache: {e}", flush=True)


def _background_render_file(fpath: str, wave_key: str, sheet_names: list) -> None:
    """Open *fpath* once and render every sheet not yet in the disk cache.

    Intended to be called from a daemon thread after a user request has
    already rendered (and returned) one sheet from this file.  The remaining
    sheets are cached to disk so the next request for those sheets is instant.
    """
    fname = os.path.basename(fpath)
    ext = fpath.lower().rsplit('.', 1)[-1]
    # Only process sheets not already on disk
    to_render = [s for s in sheet_names
                 if _load_html_from_disk(wave_key, s) is None]
    if not to_render:
        return
    try:
        if ext == 'xlsx':
            if not _OPENPYXL_OK:
                return
            wb = openpyxl.load_workbook(fpath, data_only=True)
            for sheet_name in to_render:
                try:
                    rows = _table_rows_xlsx_wb(wb, sheet_name)
                    html = _build_html(rows, wave_key, sheet_name, fname)
                    _html_cache[(wave_key, sheet_name)] = html
                    _save_html_to_disk(wave_key, sheet_name, html)
                except Exception:
                    pass
            wb.close()
        else:
            if not _XLRD_OK:
                return
            wb = xlrd.open_workbook(fpath, formatting_info=True)
            for sheet_name in to_render:
                try:
                    rows = _table_rows_xls_wb(wb, sheet_name)
                    html = _build_html(rows, wave_key, sheet_name, fname)
                    _html_cache[(wave_key, sheet_name)] = html
                    _save_html_to_disk(wave_key, sheet_name, html)
                except Exception:
                    pass
            wb.release_resources()
    except Exception:
        pass


def prerender_all_sheets():
    """Pre-render every sheet in every Vol A file into _html_cache and disk.

    Opens each file exactly ONCE and renders all its sheets in a single pass —
    much faster than opening the file once per question number.

    Disk-cached sheets are loaded from disk without re-opening the Excel file,
    so subsequent restarts are nearly instant.

    Returns (rendered_count, disk_hit_count, skipped_count).
    """
    sheet_map = get_wave_sheet_map()
    rendered = 0
    disk_hits = 0
    skipped = 0

    for wave_key, file_sheets in sheet_map.items():
        for fpath, sheet_names in file_sheets.items():
            fname = os.path.basename(fpath)
            ext = fpath.lower().rsplit('.', 1)[-1]

            # Load any sheets already on disk before opening the Excel file
            sheets_to_render = []
            for sheet_name in sheet_names:
                cache_key = (wave_key, sheet_name)
                if cache_key in _html_cache:
                    continue
                disk_html = _load_html_from_disk(wave_key, sheet_name)
                if disk_html is not None:
                    _html_cache[cache_key] = disk_html
                    disk_hits += 1
                else:
                    sheets_to_render.append(sheet_name)

            if not sheets_to_render:
                continue

            try:
                if ext == 'xlsx':
                    if not _OPENPYXL_OK:
                        continue
                    wb = openpyxl.load_workbook(fpath, data_only=True)
                    for sheet_name in sheets_to_render:
                        cache_key = (wave_key, sheet_name)
                        rows = _table_rows_xlsx_wb(wb, sheet_name)
                        html = _build_html(rows, wave_key, sheet_name, fname)
                        _html_cache[cache_key] = html
                        _save_html_to_disk(wave_key, sheet_name, html)
                        rendered += 1
                    wb.close()
                else:
                    if not _XLRD_OK:
                        continue
                    wb = xlrd.open_workbook(fpath, formatting_info=True)
                    for sheet_name in sheets_to_render:
                        cache_key = (wave_key, sheet_name)
                        rows = _table_rows_xls_wb(wb, sheet_name)
                        html = _build_html(rows, wave_key, sheet_name, fname)
                        _html_cache[cache_key] = html
                        _save_html_to_disk(wave_key, sheet_name, html)
                        rendered += 1
                    wb.release_resources()
            except Exception:
                skipped += 1

    return rendered, disk_hits, skipped


# ── Sheet lookup (O(1) from cache) ────────────────────────────────────────────

def _find_sheets_for_question(wave: str, question: str):
    """Return [(filepath, sheet_name, label), ...] by exact text matching.

    Gets the DB question text for this wave/question, normalizes it (strip QN
    prefix, collapse all whitespace to single spaces, lowercase), then scans
    every sheet in every Vol A file for the wave and returns those whose question
    cell text matches exactly after the same normalization.

    label is the question code found at the start of the matching Excel cell
    (e.g. 'QB1', 'QA3'). Falls back to the sheet name if no prefix is found.
    """
    db_text = _get_db_question_text(wave, question)
    if not db_text:
        return []
    db_norm = _normalize_exact(db_text)
    if not db_norm:
        return []

    key = _normalize_wave(wave)
    file_sheets = get_wave_sheet_map().get(key, {})

    # Batch-load uncached sheets (open each file only once)
    for fpath, sheets in file_sheets.items():
        uncached = [s for s in sheets if (fpath, s) not in _match_cache]
        if uncached:
            _batch_load_match_for_file(fpath, uncached)

    results = []
    for fpath, sheets in file_sheets.items():
        for sname in sheets:
            entry = _match_cache.get((fpath, sname))
            if entry:
                label, sheet_norm = entry
                # Accept if texts are equal, or if one is a prefix of the other
                # (Excel cells sometimes append/omit brief parenthetical instructions).
                if sheet_norm and db_norm and (
                    sheet_norm == db_norm
                    or (len(db_norm) >= 30 and sheet_norm.startswith(db_norm))
                    or (len(sheet_norm) >= 30 and db_norm.startswith(sheet_norm))
                ):
                    results.append((fpath, sname, label or sname))

    return results


def find_file_for_question(wave: str, question: str):
    """Return (filepath, is_xlsx) — backward-compat wrapper."""
    matches = _find_sheets_for_question(wave, question)
    if matches:
        fpath = matches[0][0]  # (fpath, sheet, label)[0]
        return fpath, fpath.lower().endswith('.xlsx')
    return None, None


# ── Rendering ─────────────────────────────────────────────────────────────────

def render_sheet_as_html(wave: str, question: str) -> str:
    key = _normalize_wave(wave)
    cache_key = (key, question)
    if cache_key in _html_cache:
        return _html_cache[cache_key]

    # Manual overrides — checked before any automatic matching
    ov = _overrides.get(key, {})
    if question in ov:
        entry = ov[question]
        if entry is None:
            # Explicitly no Vol A sheet for this question → TOC
            file_sheets = get_wave_sheet_map().get(key, {})
            return _toc_html(wave, question, file_sheets)
        fpath, sheet = entry[0], entry[1]
        # Overrides specify an exact (file, sheet) pair — disk cache is safe to use.
        disk_html = _load_html_from_disk(key, sheet)
        if disk_html is not None:
            _html_cache[cache_key] = disk_html
            return disk_html
        html = (_render_xlsx(fpath, sheet, wave, None) if fpath.lower().endswith('.xlsx')
                else _render_xls(fpath, sheet, wave, None))
        _html_cache[cache_key] = html
        _save_html_to_disk(key, sheet, html)
        # Background: render remaining sheets from the same file
        file_sheets = get_wave_sheet_map().get(key, {})
        all_sheets = file_sheets.get(fpath, [])
        if len(all_sheets) > 1:
            t = threading.Thread(
                target=_background_render_file,
                args=(fpath, key, all_sheets),
                daemon=True,
            )
            t.start()
        return html

    file_sheets = get_wave_sheet_map().get(key, {})

    if not file_sheets:
        # Wave not in sheet map (cache may be stale). Fall back to overrides TOC.
        wave_ov = _overrides.get(key, {})
        if wave_ov:
            synth = {}
            for _qn, _entry in wave_ov.items():
                if _entry and isinstance(_entry, (list, tuple)) and len(_entry) >= 2:
                    _fp, _sn = _entry[0], _entry[1]
                    synth.setdefault(_fp, [])
                    if _sn not in synth[_fp]:
                        synth[_fp].append(_sn)
            if synth:
                return _toc_html(wave, question, synth)
        return _error_html(
            f"Volume A file is not yet available for wave {wave}."
        )

    matches = _find_sheets_for_question(wave, question)

    # No text match — try resolving via the question index before falling back
    # to the TOC.  This handles TOC clicks where the question label (QB1) is
    # not in DuckDB but IS in the Excel file's index (QB1 → T3 for T-format,
    # or QB1 → QB1 for EB95+ where sheet name = question label).
    if not matches:
        qidx = get_question_index().get(key, {})
        idx_entry = qidx.get(question)  # [filepath, sheet_name]
        if idx_entry:
            fpath, sname = idx_entry[0], idx_entry[1]
            # Check disk cache using the actual sheet name
            disk_html = _load_html_from_disk(key, sname)
            if disk_html is not None:
                _html_cache[cache_key] = disk_html
                return disk_html
            if fpath.lower().endswith('.xlsx'):
                html = _render_xlsx(fpath, sname, wave, None)
            else:
                html = _render_xls(fpath, sname, wave, None)
            _html_cache[cache_key] = html
            _save_html_to_disk(key, sname, html)
            # Background: render remaining sheets from this file
            all_sheets = file_sheets.get(fpath, [])
            if len(all_sheets) > 1:
                t = threading.Thread(
                    target=_background_render_file,
                    args=(fpath, key, all_sheets),
                    daemon=True,
                )
                t.start()
            return html
        # Also try: question label IS a sheet name in one of the wave's files
        # (direct EB95+ hit that somehow wasn't in the question index)
        for fpath, sheets in file_sheets.items():
            if question in sheets:
                disk_html = _load_html_from_disk(key, question)
                if disk_html is not None:
                    _html_cache[cache_key] = disk_html
                    return disk_html
                if fpath.lower().endswith('.xlsx'):
                    html = _render_xlsx(fpath, question, wave, None)
                else:
                    html = _render_xls(fpath, question, wave, None)
                _html_cache[cache_key] = html
                _save_html_to_disk(key, question, html)
                # Background: render remaining sheets from this file
                if len(sheets) > 1:
                    t = threading.Thread(
                        target=_background_render_file,
                        args=(fpath, key, sheets),
                        daemon=True,
                    )
                    t.start()
                return html
        # Prefix match: QA1 matches QA1_1, QA1_2, … or QA1.1, QA1.2, …
        # Only accepted if the label extracted from the sheet content also
        # starts with the question name (e.g. "QA1.1" starts with "QA1").
        prefix = question + '_'
        prefix_dot = question + '.'
        for fpath, sheets in file_sheets.items():
            sub_sheets = [s for s in sheets
                          if s.startswith(prefix) or s.startswith(prefix_dot)]
            if not sub_sheets:
                continue
            # Verify via label extracted from cell content
            uncached = [s for s in sub_sheets if (fpath, s) not in _match_cache]
            if uncached:
                _batch_load_match_for_file(fpath, uncached)
            verified = [s for s in sub_sheets
                        if _match_cache.get((fpath, s)) and
                        _match_cache[(fpath, s)][0].upper().startswith(question.upper())]
            if not verified:
                continue
            fname = os.path.basename(fpath)
            is_xlsx = fpath.lower().endswith('.xlsx')
            sections = []
            for sname in verified:
                rows = (_table_rows_xlsx(fpath, sname) if is_xlsx
                        else _table_rows_xls(fpath, sname))
                sections.append((sname, rows))
            html = _build_html_multi(sections, wave, question, fname)
            _html_cache[cache_key] = html
            _save_html_to_disk(key, question, html)
            return html
        return _toc_html(wave, question, file_sheets)

    # Single sheet — check disk cache first, then render from Excel
    if len(matches) == 1:
        fpath, sheet, label = matches[0]
        disk_html = _load_html_from_disk(key, sheet)
        if disk_html is not None:
            _html_cache[cache_key] = disk_html
            return disk_html
        if fpath.lower().endswith('.xlsx'):
            html = _render_xlsx(fpath, sheet, wave, None)
        else:
            html = _render_xls(fpath, sheet, wave, None)
        _save_html_to_disk(key, sheet, html)
        # Background: render remaining sheets from this file
        all_sheets = file_sheets.get(fpath, [])
        if len(all_sheets) > 1:
            t = threading.Thread(
                target=_background_render_file,
                args=(fpath, key, all_sheets),
                daemon=True,
            )
            t.start()
    else:
        # Multiple matching sheets — render each as its own section with its own chart
        main_filename = os.path.basename(matches[0][0])
        sections = []
        for fpath, sheet, label in matches:
            rows = (_table_rows_xlsx(fpath, sheet) if fpath.lower().endswith('.xlsx')
                    else _table_rows_xls(fpath, sheet))
            sections.append((label, rows))
        html = _build_html_multi(sections, wave, question, main_filename)

    _html_cache[cache_key] = html
    return html


# ── Whole-file renderers (open once, render all sheets) ──────────────────────

def _all_sheets_xls(filepath, sheet_names) -> str:
    """Open .xls once and render all requested sheets."""
    try:
        wb = xlrd.open_workbook(filepath, formatting_info=True)
    except Exception as e:
        return f'<tr><td style="color:red">Could not open file: {_esc(str(e))}</td></tr>'
    combined = ''
    for sheet_name in sheet_names:
        combined += (
            f'<tr><td colspan="30" style="background:#2c5f8a;color:#fff;'
            f'font-weight:bold;padding:4px 8px;font-size:12px;">'
            f'Sheet: {_esc(sheet_name)}</td></tr>\n'
        )
        combined += _table_rows_xls_wb(wb, sheet_name)
    wb.release_resources()
    return combined


def _all_sheets_xlsx(filepath, sheet_names) -> str:
    """Open .xlsx once and render all requested sheets."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        return f'<tr><td style="color:red">Could not open file: {_esc(str(e))}</td></tr>'
    combined = ''
    for sheet_name in sheet_names:
        combined += (
            f'<tr><td colspan="30" style="background:#2c5f8a;color:#fff;'
            f'font-weight:bold;padding:4px 8px;font-size:12px;">'
            f'Sheet: {_esc(sheet_name)}</td></tr>\n'
        )
        combined += _table_rows_xlsx_wb(wb, sheet_name)
    wb.close()
    return combined


# ── .xls renderer (xlrd) ─────────────────────────────────────────────────────

def _pct_format(v, fmt_str: str) -> str:
    """Format a decimal value as a percentage matching Excel's format string."""
    v_pct = v * 100
    m = re.search(r'\.(\d+)%', fmt_str)
    decimals = len(m.group(1)) if m else 0
    return f'{v_pct:.{decimals}f}%'


# xlrd built-in percentage format IDs (not stored in wb.format_map)
_XLRD_BUILTIN_PCT = {9: '0%', 10: '0.00%'}


def _fmt_xlrd_value(cell, wb):
    """Format an xlrd cell value as a string, respecting number formats."""
    import xlrd as _xlrd
    if cell.ctype in (_xlrd.XL_CELL_EMPTY, _xlrd.XL_CELL_BLANK):
        return ''
    if cell.ctype == _xlrd.XL_CELL_ERROR:
        return ''
    if cell.ctype == _xlrd.XL_CELL_NUMBER:
        v = cell.value
        try:
            fmt_key = wb.xf_list[cell.xf_index].format_key
            # Built-in % formats (9=0%, 10=0.00%) are NOT in wb.format_map
            if fmt_key in _XLRD_BUILTIN_PCT:
                return _pct_format(v, _XLRD_BUILTIN_PCT[fmt_key])
            # Custom formats defined in the file
            fmt_obj = wb.format_map.get(fmt_key)
            if fmt_obj and '%' in fmt_obj.format_str:
                return _pct_format(v, fmt_obj.format_str)
        except Exception:
            pass
        if v == int(v):
            return str(int(v))
        return f'{v:.4g}'
    return str(cell.value)


def _question_row_range_xls(ws, question: str):
    """Scan an xlrd sheet's col1 for question headers; return (start, end) row indices.

    Finds the first row where col1 starts with '<question> ' and the first row
    after that where a *different* question number begins.  Returns None if the
    question header is not found in col1 at all (e.g. T-format sheets where the
    whole sheet belongs to one question).
    """
    qn_rows = []   # [(row_index, matched_question_number)]
    for r in range(ws.nrows):
        try:
            cell_val = str(ws.cell(r, 1).value).strip()
        except Exception:
            continue
        m = _QN_RE.match(cell_val)
        if m:
            qn_rows.append((r, m.group(1)))

    if not qn_rows:
        return None

    start = next((r for r, qn in qn_rows if qn == question), None)
    if start is None:
        return None

    end = ws.nrows
    for r, qn in qn_rows:
        if r > start and qn != question:
            end = r
            break

    return (start, end)


def _table_rows_xls_wb(wb, sheet_name, row_start=0, row_end=None) -> str:
    """Extract table rows HTML from an already-open xlrd workbook.

    row_start / row_end limit rendering to a slice of rows (used when a single
    sheet contains multiple questions and we only want one of them).
    """
    if sheet_name not in wb.sheet_names():
        return f'<tr><td style="color:red">Sheet {_esc(sheet_name)} not found.</td></tr>'

    ws = wb.sheet_by_name(sheet_name)
    if row_end is None:
        row_end = ws.nrows

    # Column widths: xlrd stores in 1/256 char units; 1 char ≈ 12px at 11px font
    col_widths = []
    for c in range(ws.ncols):
        try:
            w_px = max(80, min(600, int(ws.col(c).width / 256 * 12)))
        except Exception:
            w_px = 80
        col_widths.append(w_px)

    skip = set()
    merge_info = {}
    for r1, r2, c1, c2 in ws.merged_cells:
        merge_info[(r1, c1)] = (r2 - r1, c2 - c1)
        for r in range(r1, r2):
            for c in range(c1, c2):
                if not (r == r1 and c == c1):
                    skip.add((r, c))

    table_rows = ''
    for r in range(row_start, row_end):
        row_html = ''
        row_has_content = False

        # Col A → Col B shift: if col A has text and col B is empty, move the
        # label to col B so Total/header rows align with answer rows (same fix
        # already applied in the xlsx renderer).
        if ws.ncols >= 2:
            v0 = _fmt_xlrd_value(ws.cell(r, 0), wb)
            v1 = _fmt_xlrd_value(ws.cell(r, 1), wb)
            v0 = '' if v0.lstrip('<').lower().startswith('back to content') else v0
            v1 = '' if v1.lstrip('<').lower().startswith('back to content') else v1
            if v0.strip() and not v1.strip():
                _col_override = {0: '', 1: v0}
            else:
                _col_override = {0: v0, 1: v1}
        else:
            _col_override = {}

        for c in range(ws.ncols):
            if (r, c) in skip:
                continue
            cell = ws.cell(r, c)
            if c in _col_override:
                val = _col_override[c]
            else:
                val = _fmt_xlrd_value(cell, wb)
                if val.lstrip('<').lower().startswith('back to content'):
                    val = ''
            if val:
                row_has_content = True
            span_attrs = ''
            if (r, c) in merge_info:
                rs, cs = merge_info[(r, c)]
                if rs > 1:
                    span_attrs += f' rowspan="{rs}"'
                if cs > 1:
                    span_attrs += f' colspan="{cs}"'
            w_px = col_widths[c] if c < len(col_widths) else 80
            style = _cell_style(c, val) + f';min-width:{w_px}px'
            row_html += f'<td style="{style}"{span_attrs}>{_esc(val)}</td>'
        if row_has_content:
            try:
                h_twips = ws.row(r).height  # 1/20 of a point
                tr_style = f' style="height:{max(15, int(h_twips / 20 * 1.333))}px"' if h_twips > 0 else ''
            except Exception:
                tr_style = ''
            table_rows += f'<tr{tr_style}>{row_html}</tr>\n'
        else:
            # Render invisible placeholder to keep rowspan counts accurate.
            # Skipping empty rows shifts rowspan cells beyond their intended range.
            table_rows += '<tr style="height:0;line-height:0;font-size:0"></tr>\n'

    return table_rows


def _table_rows_xls(filepath, sheet_name, question_filter=None) -> str:
    """Extract table rows HTML from an .xls sheet (no full-page wrapper).

    If question_filter is given, only rows belonging to that question are rendered
    (detected by scanning col1 for question-number headers).
    """
    try:
        wb = xlrd.open_workbook(filepath, formatting_info=True)
    except Exception as e:
        return f'<tr><td style="color:red">Could not open file: {_esc(str(e))}</td></tr>'

    row_start, row_end = 0, None
    if question_filter:
        ws = wb.sheet_by_name(sheet_name) if sheet_name in wb.sheet_names() else None
        if ws:
            rng = _question_row_range_xls(ws, question_filter)
            if rng:
                row_start, row_end = rng

    result = _table_rows_xls_wb(wb, sheet_name, row_start, row_end)
    wb.release_resources()
    return result


def _render_xls(filepath, sheet_name, wave, question_filter=None):
    rows = _table_rows_xls(filepath, sheet_name, question_filter)
    label = question_filter if question_filter else sheet_name
    return _build_html(rows, wave, label, os.path.basename(filepath))


# ── .xlsx renderer (openpyxl) ─────────────────────────────────────────────────

def _color_to_hex(color_obj):
    """openpyxl Color → '#RRGGBB' or None."""
    if color_obj is None:
        return None
    try:
        if color_obj.type == 'rgb':
            argb = color_obj.rgb  # 8-char ARGB
            if len(argb) == 8 and argb[:2] != '00':
                return f'#{argb[2:]}'
    except Exception:
        pass
    return None


def _question_row_range_xlsx(ws, question: str):
    """Scan an openpyxl sheet's col B (index 2) for question headers; return (start, end).

    Row indices are 0-based (converted from openpyxl's 1-based).
    Returns None if question header not found.
    """
    qn_rows = []
    for row in ws.iter_rows(min_col=2, max_col=2, values_only=True):
        cell_val = row[0]
        if cell_val is None:
            qn_rows.append(None)
            continue
        m = _QN_RE.match(str(cell_val).strip())
        qn_rows.append(m.group(1) if m else None)

    start = next((i for i, qn in enumerate(qn_rows) if qn == question), None)
    if start is None:
        return None

    end = len(qn_rows)
    for i in range(start + 1, len(qn_rows)):
        if qn_rows[i] and qn_rows[i] != question:
            end = i
            break

    return (start, end)


def _table_rows_xlsx_wb(wb, sheet_name, row_start=0, row_end=None) -> str:
    """Extract table rows HTML from an already-open openpyxl workbook."""
    if sheet_name not in wb.sheetnames:
        return f'<tr><td style="color:red">Sheet {_esc(sheet_name)} not found.</td></tr>'

    ws = wb[sheet_name]

    # Column widths: openpyxl stores in char units; 1 char ≈ 12px at 11px font
    from openpyxl.utils import get_column_letter
    max_col = ws.max_column or 0
    col_widths = {}
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        dim = ws.column_dimensions.get(letter)
        try:
            w_px = max(80, min(600, int((dim.width or 8) * 12))) if dim else 100
        except Exception:
            w_px = 80
        col_widths[c] = w_px

    skip = set()
    merge_info = {}
    for rng in ws.merged_cells.ranges:
        r1, c1, r2, c2 = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        merge_info[(r1, c1)] = (r2 - r1 + 1, c2 - c1 + 1)
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                if not (r == r1 and c == c1):
                    skip.add((r, c))

    # Convert 0-based row_start/row_end to openpyxl 1-based
    min_row = (row_start or 0) + 1
    max_row = row_end  # None = no limit; openpyxl accepts None for max_row

    table_rows = ''
    for row in ws.iter_rows(min_row=min_row, max_row=max_row):
        row_html = ''
        row_has_content = False
        row_num = row[0].row if row else None

        # Alignment fix: if col A (index 0) has text and col B (index 1) is
        # empty, shift col A's value into col B so Total rows align with
        # answer rows that always have their label in col B.
        row_str = [str(c.value or '') for c in row]
        # Strip nav artifacts before deciding whether to shift
        row_str = ['' if v.lstrip('<').lower().startswith('back to content') else v
                   for v in row_str]
        if (len(row_str) >= 2
                and row_str[0].strip()
                and not row_str[1].strip()):
            row_str[1] = row_str[0]
            row_str[0] = ''
        _label_override = row_str  # only used for cols 0 and 1

        for cell in row:
            if (cell.row, cell.column) in skip:
                continue

            col_idx_0 = cell.column - 1  # 0-based
            # For label columns (A and B), use the pre-shifted, pre-filtered value
            if col_idx_0 <= 1 and col_idx_0 < len(_label_override):
                val = _label_override[col_idx_0]
            else:
                val = cell.value
                if val is None:
                    val = ''
                elif isinstance(val, (int, float)):
                    fmt_str = cell.number_format or ''
                    if '%' in fmt_str:
                        val = _pct_format(float(val), fmt_str)
                    elif isinstance(val, float):
                        val = f'{val:.4g}' if val != int(val) else str(int(val))
                    else:
                        val = str(val)
                else:
                    val = str(val)
                if val.lstrip('<').lower().startswith('back to content'):
                    val = ''

            if val:
                row_has_content = True

            parts = []
            try:
                if cell.fill and cell.fill.fill_type == 'solid':
                    bg = _color_to_hex(cell.fill.fgColor)
                    if bg:
                        parts.append(f'background-color:{bg}')
            except Exception:
                pass
            try:
                if cell.font:
                    if cell.font.bold:
                        parts.append('font-weight:bold')
                    if cell.font.italic:
                        parts.append('font-style:italic')
                    fc = _color_to_hex(cell.font.color)
                    if fc:
                        parts.append(f'color:{fc}')
            except Exception:
                pass
            try:
                if cell.alignment:
                    ha = cell.alignment.horizontal
                    if ha in ('left', 'right', 'center', 'justify'):
                        parts.append(f'text-align:{ha}')
            except Exception:
                pass

            w_px = col_widths.get(cell.column, 80)
            if not parts:
                parts = [_cell_style(cell.column - 1, val) + f';min-width:{w_px}px']
            else:
                parts.insert(0, 'padding:3px 8px;border:1px solid #ccc')
                parts.append(f'min-width:{w_px}px')
                if not any('text-align' in p for p in parts):
                    ta = 'text-align:left' if col_idx_0 <= 1 else 'text-align:center'
                    parts.append(ta)

            span_attrs = ''
            if (cell.row, cell.column) in merge_info:
                rs, cs = merge_info[(cell.row, cell.column)]
                if rs > 1:
                    span_attrs += f' rowspan="{rs}"'
                if cs > 1:
                    span_attrs += f' colspan="{cs}"'

            row_html += f'<td style="{";".join(parts)}"{span_attrs}>{_esc(val)}</td>'

        if row_has_content:
            tr_style = ''
            if row_num:
                try:
                    rd = ws.row_dimensions.get(row_num)
                    if rd and rd.height:
                        h_px = max(15, int(rd.height * 1.333))
                        tr_style = f' style="height:{h_px}px"'
                except Exception:
                    pass
            table_rows += f'<tr{tr_style}>{row_html}</tr>\n'
        else:
            # Render invisible placeholder to keep rowspan counts accurate.
            # Skipping empty rows shifts rowspan cells beyond their intended range.
            table_rows += '<tr style="height:0;line-height:0;font-size:0"></tr>\n'

    return table_rows


def _table_rows_xlsx(filepath, sheet_name, question_filter=None) -> str:
    """Extract table rows HTML from an .xlsx sheet (no full-page wrapper)."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        return f'<tr><td style="color:red">Could not open file: {_esc(str(e))}</td></tr>'

    row_start, row_end = 0, None
    if question_filter and sheet_name in wb.sheetnames:
        rng = _question_row_range_xlsx(wb[sheet_name], question_filter)
        if rng:
            row_start, row_end = rng

    result = _table_rows_xlsx_wb(wb, sheet_name, row_start, row_end)
    wb.close()
    return result


def _render_xlsx(filepath, sheet_name, wave, question_filter=None):
    rows = _table_rows_xlsx(filepath, sheet_name, question_filter)
    label = question_filter if question_filter else sheet_name
    return _build_html(rows, wave, label, os.path.basename(filepath))


# ── Shared HTML helpers ───────────────────────────────────────────────────────

def _cell_style(col_idx: int, val: str) -> str:
    """Positional default style used when no Excel style info is available."""
    base = 'padding:3px 8px;border:1px solid #ccc;'
    if col_idx <= 1:  # cols A and B are label columns
        return base + 'text-align:left'
    return base + 'text-align:center'


_CHART_JS = """
<script src="https://cdn.jsdelivr.net/npm/chart.js@4/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2/dist/chartjs-plugin-datalabels.min.js"></script>
<script>
Chart.register(ChartDataLabels);
var PALETTE=['#2c5f8a','#27ae60','#e67e22','#c0392b','#95a5a6','#3d7ab5','#82e0aa','#f39c12','#8e44ad','#d35400','#1abc9c','#e74c3c'];

/* Parse cell text as integer percentage (0-100), or null if it's a count/invalid.
   "7%" -> 7,  "0.07" -> 7,  "7" -> null (ambiguous count),  "665" -> null */
function parsePct(text){
  var t=(text||'').trim();
  if(!t||t==='-'||t==='*')return null;
  if(t.charAt(t.length-1)==='%'){
    var n=parseFloat(t);
    return isNaN(n)?null:Math.round(n);
  }
  var n=parseFloat(t);
  if(isNaN(n)||n<0)return null;
  if(t.indexOf('.')!==-1&&n<=1) return Math.round(n*100); // decimal proportion e.g. 0.07
  return null; // integer without % = count
}

/* Content-based parser: finds "pct rows" by scanning for cells ending with "%".
   This avoids all rowspan/colspan column-index issues — we locate data by
   content, not by physical or logical column position. */
function parseVolATable(tblId){
  var tbl=document.getElementById(tblId);
  if(!tbl)return null;
  var rows=Array.from(tbl.querySelectorAll('tr'));

  // 1. Find the header row containing EU27/UE27/EU28/UE28 column label
  var headerRowIdx=-1, eu27PhysIdx=-1, colNames=[];
  for(var i=0;i<rows.length;i++){
    var cells=Array.from(rows[i].querySelectorAll('td'));
    for(var j=0;j<cells.length;j++){
      if(/EU\\s*\\d+|UE\\s*\\d+/i.test(cells[j].textContent.trim())){
        headerRowIdx=i; eu27PhysIdx=j;
        // Collect column names: EU aggregate + all non-empty cells after it in this row
        colNames=[cells[j].textContent.trim()];
        for(var k=j+1;k<cells.length;k++){
          var nm=cells[k].textContent.trim();
          if(nm) colNames.push(nm);
        }
        break;
      }
    }
    if(headerRowIdx!==-1)break;
  }
  if(headerRowIdx===-1||!colNames.length)return null;

  // 2. Scan data rows: a "pct row" is one where we can find %-ending cells.
  //    We locate the start of the data section by finding the first cell with "%".
  //    Everything before it in the same row = label candidates.
  var answers=[], seenLabels={};
  for(var r=headerRowIdx+1;r<rows.length;r++){
    var cells=Array.from(rows[r].querySelectorAll('td'));
    if(!cells.length)continue;

    // Find first cell ending with "%" — that is the EU27 pct value
    var dataStart=-1;
    for(var j=0;j<cells.length;j++){
      var t=cells[j].textContent.trim();
      if(t.charAt(t.length-1)==='%'){dataStart=j;break;}
    }
    if(dataStart===-1)continue; // count row or empty — skip

    // Label: first non-empty, non-numeric cell BEFORE the data section
    var label='';
    for(var j=0;j<dataStart;j++){
      var t=cells[j].textContent.trim();
      if(t&&!/^[\\d\\s\\-]+$/.test(t)){label=t;break;}
    }
    // If label is empty (rowspan covers it), look at the previous row's label
    if(!label&&r>headerRowIdx+1){
      var prev=Array.from(rows[r-1].querySelectorAll('td'));
      for(var j=0;j<prev.length;j++){
        var t=prev[j].textContent.trim();
        if(t&&!/^[\\d\\s\\-]+$/.test(t)&&!/^[0-9.]+%$/.test(t)){label=t;break;}
      }
    }
    if(!label||/^total/i.test(label))continue;
    if(seenLabels[label])continue; // deduplicate (summary section repeats rows)
    seenLabels[label]=1;

    // Collect data values starting from dataStart; pad with null for missing cells
    var dataVals=[];
    for(var j=dataStart;j<cells.length;j++){
      dataVals.push(parsePct(cells[j].textContent.trim()));
    }
    var eu27Pct=dataVals[0];
    if(eu27Pct===null)continue;

    // Map country values: colNames[1..] -> dataVals[1..]
    var cVals={};
    for(var ci=1;ci<colNames.length&&ci<dataVals.length+1;ci++){
      cVals[colNames[ci]]=dataVals[ci] !== undefined ? dataVals[ci] : null;
    }
    answers.push({label:label,eu27:eu27Pct,cVals:cVals});
  }
  if(!answers.length)return null;
  var countries=colNames.slice(1).filter(function(c){return !/EU\\s*27|UE\\s*27/i.test(c);});
  return{
    answerLabels:answers.map(function(a){return a.label;}),
    eu27Data:answers.map(function(a){return a.eu27;}),
    countries:countries,
    datasets:answers.map(function(a,i){return{
      label:a.label,
      data:countries.map(function(c){var v=a.cVals[c];return(v!==null&&v!==undefined)?v:0;}),
      backgroundColor:PALETTE[i%PALETTE.length]
    };})
  };
}

function renderVolACharts(tblId,pieId,barId){
  var d=parseVolATable(tblId);
  var pieCanvas=document.getElementById(pieId);
  var barCanvas=document.getElementById(barId);
  if(!d||!d.answerLabels.length){
    if(pieCanvas)pieCanvas.closest('.chart-section').style.display='none';
    if(barCanvas)barCanvas.closest('.chart-section').style.display='none';
    return;
  }
  if(pieCanvas){
    new Chart(pieCanvas,{
      type:'pie',
      data:{
        labels:d.answerLabels,
        datasets:[{data:d.eu27Data,backgroundColor:PALETTE.slice(0,d.answerLabels.length)}]
      },
      options:{responsive:true,maintainAspectRatio:true,
        plugins:{
          legend:{position:'right',labels:{font:{size:10},boxWidth:12}},
          tooltip:{callbacks:{label:function(c){return c.label+': '+c.parsed+'%';}}},
          datalabels:{color:'#fff',font:{size:10,weight:'bold'},
            formatter:function(v){return v>4?v+'%':'';}}
        }
      }
    });
  }
  if(barCanvas&&d.countries.length){
    var barWrap=barCanvas.parentElement;
    var h=Math.max(180,d.countries.length*15)+'px';
    barWrap.style.height=h;
    barCanvas.style.height=h;
    new Chart(barCanvas,{
      type:'bar',
      data:{labels:d.countries,datasets:d.datasets},
      options:{
        indexAxis:'y',responsive:true,maintainAspectRatio:false,animation:false,
        scales:{
          x:{stacked:true,max:100,ticks:{callback:function(v){return v+'%';},font:{size:8}}},
          y:{stacked:true,ticks:{font:{size:8}}}
        },
        plugins:{
          legend:{position:'top',labels:{font:{size:9},boxWidth:10}},
          tooltip:{callbacks:{label:function(c){return c.dataset.label+': '+c.parsed.x+'%';}}},
          datalabels:{color:'#fff',font:{size:8},anchor:'center',align:'center',
            formatter:function(v){return v>6?v+'%':'';}}
        }
      }
    });
  }
}
</script>
"""

_CHART_CSS = """
    .chart-wrap { margin: 8px 0 32px; display: none; }
    .chart-section { margin-bottom: 20px; }
    .chart-label { color: #2c5f8a; font-size: 11px; font-weight: bold; margin: 8px 0 4px; }
    .chart-btn { margin: 6px 0; padding: 4px 14px; font-size: 11px; cursor: pointer; background: #2c5f8a; color: #fff; border: none; border-radius: 3px; }
    .chart-btn:hover { background: #3d7ab5; }
"""

_CHART_TOGGLE_JS = """
<script>
var _chartDone={};
function toggleChart(btn,tblId,wrapId){
  var w=document.getElementById(wrapId);
  if(w.style.display==='block'){w.style.display='none';btn.textContent='Show charts';}
  else{w.style.display='block';btn.textContent='Hide charts';
    if(!_chartDone[wrapId]){
      var n=wrapId.replace('cwrap-','');
      renderVolACharts(tblId,'pie-'+n,'bar-'+n);
      _chartDone[wrapId]=1;
    }
  }
}
window.addEventListener('load',function(){
  if(new URLSearchParams(location.search).get('charts')==='1'){
    document.querySelectorAll('.chart-btn').forEach(function(btn){btn.click();});
  }
});
</script>
"""

def _chart_block(i):
    """Return the toggle button + chart containers for table index i."""
    return (
        f'<button class="chart-btn" onclick="toggleChart(this,\'tbl-{i}\',\'cwrap-{i}\')">Show charts</button>'
        f'<div class="chart-wrap" id="cwrap-{i}">'
        f'  <div class="chart-section"><p class="chart-label">EU27 aggregate</p>'
        f'  <div style="max-width:480px"><canvas id="pie-{i}"></canvas></div></div>'
        f'  <div class="chart-section"><p class="chart-label">By country</p>'
        f'  <div id="barwrap-{i}" style="overflow:hidden"><canvas id="bar-{i}"></canvas></div></div>'
        f'</div>'
    )

_PAGE_STYLE = """
    body { font-family: Arial, sans-serif; font-size: 11px; padding: 16px; color: #222; background: #fff; }
    h2 { color: #2c5f8a; margin-bottom: 4px; }
    h3 { color: #2c5f8a; margin: 24px 0 4px; }
    .meta { color: #666; font-size: 10px; margin-bottom: 12px; }
    .scroll-wrap { overflow-x: auto; }
    table { border-collapse: collapse; }
    td { vertical-align: top; white-space: pre-wrap; word-break: break-word; }
    td[rowspan] { vertical-align: middle; }
"""

def _build_html(table_rows, wave, question, filename):
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8"/>
  <title>Volume A — Wave {_esc(wave)} / {_esc(question)}</title>
  {_CHART_JS}
  {_CHART_TOGGLE_JS}
  <style>{_PAGE_STYLE}{_CHART_CSS}</style>
</head>
<body>
  <h2>Volume A — Wave {_esc(wave)}, Question {_esc(question)}</h2>
  <p class="meta">Source: {_esc(filename)}</p>
  <div class="scroll-wrap"><table id="tbl-0">{table_rows}</table></div>
  {_chart_block(0)}
</body>
</html>"""


def _build_html_multi(sections, wave, question, filename):
    """Render multiple sub-tables (one per sub-sheet) with optional per-section charts."""
    sections_html = ''
    for i, (sname, table_rows) in enumerate(sections):
        sections_html += (
            f'<h3>{_esc(sname)}</h3>'
            f'<div class="scroll-wrap"><table id="tbl-{i}">{table_rows}</table></div>'
            + _chart_block(i) + '\n'
        )
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8"/>
  <title>Volume A — Wave {_esc(wave)} / {_esc(question)}</title>
  {_CHART_JS}
  {_CHART_TOGGLE_JS}
  <style>{_PAGE_STYLE}{_CHART_CSS}</style>
</head>
<body>
  <h2>Volume A — Wave {_esc(wave)}, Question {_esc(question)}</h2>
  <p class="meta">Source: {_esc(filename)}</p>
  {sections_html}
</body>
</html>"""


def _build_html_with_notice(table_rows, wave, question, filename, notice):
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8"/>
  <title>Volume A — Wave {_esc(wave)} / {_esc(question)}</title>
  <style>
    body {{
      font-family: Arial, sans-serif;
      font-size: 11px;
      padding: 16px;
      color: #222;
      background: #fff;
    }}
    h2 {{ color: #2c5f8a; margin-bottom: 4px; }}
    .meta {{ color: #666; font-size: 10px; margin-bottom: 12px; }}
    .scroll-wrap {{ overflow-x: auto; }}
    table {{ border-collapse: collapse; }}
    td {{ vertical-align: top; white-space: pre-wrap; word-break: break-word; }}
    td[rowspan] {{ vertical-align: middle; }}
  </style>
</head>
<body>
  <h2>Volume A — Wave {_esc(wave)}, Question {_esc(question)}</h2>
  <p class="meta">Source: {_esc(filename)}</p>
  {notice}
  <div class="scroll-wrap">
    <table>{table_rows}</table>
  </div>
</body>
</html>"""


def _toc_html(wave: str, question: str, file_sheets: dict) -> str:
    """Return a fast table-of-contents page when no sheet matches question.

    Each entry links to /api/volume-a?wave=...&question=SHEETNAME so only
    the selected sheet is rendered on click.  Displays the question label
    (e.g. QB1, QA3) from the question index instead of raw sheet names.
    """
    import urllib.parse
    wave_enc = urllib.parse.quote(wave)

    # Build reverse mapping: sheet_name → question_label, from question index.
    # For T-format waves: T3 → QB1. For new-format: QB1 → QB1 (identity).
    key = _normalize_wave(wave)
    qidx = get_question_index().get(key, {})
    sheet_to_label = {entry[1]: qn for qn, entry in qidx.items()}

    sections = ''
    for fpath, sheets in file_sheets.items():
        fname = _esc(os.path.basename(fpath))
        links = ''
        for s in sheets:
            label = sheet_to_label.get(s, s)  # question label (QB1, QA3…) or sheet name as fallback
            q_enc = urllib.parse.quote(label)  # link uses label so text matching works on click
            href = f'/api/volume-a?wave={wave_enc}&question={q_enc}'
            links += (
                f'<li><a href="{href}" '
                f'style="color:#2c5f8a;text-decoration:none;">'
                f'{_esc(label)}</a></li>\n'
            )
        sections += f'<h3 style="margin:16px 0 4px;color:#4a4a4a;font-size:12px;">{fname}</h3><ul style="margin:0 0 12px;padding-left:20px;line-height:1.8;">{links}</ul>\n'

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8"/>
  <title>Volume A — Wave {_esc(wave)} / {_esc(question)}</title>
  <style>
    body {{ font-family: Arial, sans-serif; font-size: 12px; padding: 16px; color: #222; background: #fff; }}
    h2 {{ color: #2c5f8a; margin-bottom: 4px; }}
    a:hover {{ text-decoration: underline !important; }}
  </style>
</head>
<body>
  <h2>Volume A — Wave {_esc(wave)}</h2>
  <p style="background:#fff3cd;border:1px solid #ffc107;padding:8px 12px;
     border-radius:4px;font-size:11px;margin-bottom:16px;">
    No sheet matching <b>{_esc(question)}</b> found.
    Click a sheet below to open it.
  </p>
  {sections}
</body>
</html>"""


def _error_html(msg):
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"/><title>Volume A — Not found</title></head>
<body style="font-family:Arial,sans-serif;padding:24px;color:#900;">
  <h3>Volume A data not available</h3><p>{msg}</p>
</body></html>"""


def _esc(text):
    return (str(text)
            .replace('&', '&amp;')
            .replace('<', '&lt;')
            .replace('>', '&gt;'))
